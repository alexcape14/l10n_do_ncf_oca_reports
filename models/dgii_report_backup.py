# -*- coding: utf-8 -*-
# © 2015-2018 Marcos Organizador de Negocios SRL. (https://marcos.do/)

#             Eneldo Serrata <eneldo@marcos.do>
# © 2018-2019 SoftNet Team SRL. (https://www.softnet.do/)
#             Manuel Gonzalez <manuel@softnet.do>

# © 2018 Jeffry Jesus De La Rosa <jeffryjesus@gmail.com>

# This file is part of NCF DGII Reports
# NCF DGII Reports is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# NCF DGII Reports is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with NCF DGII Reports.  If not, see <http://www.gnu.org/licenses/>.
# ######################################################################..

"""
DGII Report
"""
import base64
import os
import re
import calendar
import logging
# from var_dump import var_dump
from pprint import pprint as pp
from openpyxl import load_workbook

from odoo import models, fields, api, exceptions

_logger = logging.getLogger(__name__)

try:
	from stdnum.do import ncf, rnc, cedula
except(ImportError, IOError) as err:
	_logger.debug(err)


class DgiiReport(models.Model):
	_name = "dgii.report"
	_inherit = ['mail.thread', 'mail.activity.mixin']

	@api.multi
	@api.depends("purchase_report")
	def _purchase_report_totals(self):

		# Tipos de Bienes y Servicios Comprados
		# Columna 3 del 606
		summary_dict = {
			"01": {"count": 0, "amount": 0.0},
			"02": {"count": 0, "amount": 0.0},
			"03": {"count": 0, "amount": 0.0},
			"04": {"count": 0, "amount": 0.0},
			"05": {"count": 0, "amount": 0.0},
			"06": {"count": 0, "amount": 0.0},
			"07": {"count": 0, "amount": 0.0},
			"08": {"count": 0, "amount": 0.0},
			"09": {"count": 0, "amount": 0.0},
			"10": {"count": 0, "amount": 0.0},
			"11": {"count": 0, "amount": 0.0},
		}

		for rec in self:  # self  = lines on model DgiiReportPurchaseLine

			if rec.name:
				report_month, report_year = rec.name.split("/")
			else:
				report_month = report_year = False

			rec.ITBIS_TOTAL = 0
			rec.ITBIS_TOTAL_NC = 0
			rec.ITBIS_TOTAL_PAYMENT = 0
			rec.TOTAL_MONTO_FACTURADO = 0
			rec.MONTO_FACTURADO_SERVICIOS = 0
			rec.MONTO_FACTURADO_BIENES = 0
			rec.TOTAL_MONTO_NC = 0
			rec.TOTAL_MONTO_PAYMENT = 0
			rec.ITBIS_RETENIDO = 0
			rec.RETENCION_RENTA = 0
			rec.ITBIS_FACTURADO_BIENES = 0
			rec.ITBIS_FACTURADO_SERVICIOS = 0
			rec.ITBIS_SUJETO_PROPORCIONALIDAD = 0
			rec.ITBIS_LLEVADO_ALCOSTO = 0
			rec.ITBIS_POR_ADELANTAR = 0
			rec.ITBIS_PERCIBIDO_COMPRAS = 0
			rec.RETENCION_RENTA = 0
			rec.ISR_PERCIBIDO_COMPRAS = 0
			rec.IMPUESTO_ISC = 0
			rec.IMPUESTOS_OTROS = 0
			rec.MONTO_PROPINA_LEGAL = 0

			for purchase in rec.purchase_report:

				TIPO_COMPROBANTE = self.get_type_voucher(purchase)
				ncf_year, ncf_month, ncf_day = purchase.FECHA_COMPROBANTE.split("-")

				if TIPO_COMPROBANTE == "04":  # 04 = NOTAS DE CRÉDITOS #TODO check to validate NC for Monto Facturado Bienes/Servicios

					rec.ITBIS_TOTAL_NC += purchase.ITBIS_FACTURADO_TOTAL
					rec.TOTAL_MONTO_NC += purchase.MONTO_FACTURADO
					rec.ITBIS_RETENIDO -= purchase.ITBIS_RETENIDO
					rec.ITBIS_FACTURADO_SERVICIOS -= purchase.ITBIS_FACTURADO_SERVICIOS
					rec.ITBIS_FACTURADO_BIENES -= purchase.ITBIS_FACTURADO_BIENES
					rec.ITBIS_SUJETO_PROPORCIONALIDAD -= purchase.ITBIS_SUJETO_PROPORCIONALIDAD
					rec.ITBIS_LLEVADO_ALCOSTO -= purchase.ITBIS_LLEVADO_ALCOSTO
					rec.ITBIS_POR_ADELANTAR -= purchase.ITBIS_POR_ADELANTAR
					rec.ITBIS_PERCIBIDO_COMPRAS -= purchase.ITBIS_PERCIBIDO_COMPRAS
					rec.RETENCION_RENTA -= purchase.RETENCION_RENTA
					rec.ISR_PERCIBIDO_COMPRAS -= purchase.ISR_PERCIBIDO_COMPRAS
					rec.IMPUESTO_ISC -= purchase.IMPUESTO_ISC
					rec.IMPUESTOS_OTROS -= purchase.IMPUESTOS_OTROS
					rec.MONTO_PROPINA_LEGAL -= purchase.MONTO_PROPINA_LEGAL

				elif purchase.NUMERO_COMPROBANTE_MODIFICADO == False:

					if int(report_month) == int(
							ncf_month) and report_year == ncf_year:  # this validation is to avoid add amounts of invoices of previous months
						rec.TOTAL_MONTO_FACTURADO += purchase.MONTO_FACTURADO
						rec.MONTO_FACTURADO_SERVICIOS += purchase.MONTO_FACTURADO_SERVICIOS
						rec.MONTO_FACTURADO_BIENES += purchase.MONTO_FACTURADO_BIENES
						rec.ITBIS_TOTAL += purchase.ITBIS_FACTURADO_TOTAL
						rec.ITBIS_FACTURADO_SERVICIOS += purchase.ITBIS_FACTURADO_SERVICIOS
						rec.ITBIS_FACTURADO_BIENES += purchase.ITBIS_FACTURADO_BIENES
						rec.ITBIS_RETENIDO += purchase.ITBIS_RETENIDO
						rec.RETENCION_RENTA += purchase.RETENCION_RENTA

					rec.ITBIS_SUJETO_PROPORCIONALIDAD += purchase.ITBIS_SUJETO_PROPORCIONALIDAD
					rec.ITBIS_LLEVADO_ALCOSTO += purchase.ITBIS_LLEVADO_ALCOSTO
					rec.ITBIS_POR_ADELANTAR += purchase.ITBIS_POR_ADELANTAR
					rec.ITBIS_PERCIBIDO_COMPRAS += purchase.ITBIS_PERCIBIDO_COMPRAS
					rec.ISR_PERCIBIDO_COMPRAS += purchase.ISR_PERCIBIDO_COMPRAS
					rec.IMPUESTO_ISC += purchase.IMPUESTO_ISC
					rec.IMPUESTOS_OTROS += purchase.IMPUESTOS_OTROS
					rec.MONTO_PROPINA_LEGAL += purchase.MONTO_PROPINA_LEGAL

				summary_dict[purchase.invoice_id.expense_type]["count"] += 1
				summary_dict[purchase.invoice_id.expense_type]["amount"] += purchase.MONTO_FACTURADO

			rec.ITBIS_TOTAL_PAYMENT = rec.ITBIS_TOTAL - rec.ITBIS_TOTAL_NC
			rec.TOTAL_MONTO_PAYMENT = rec.TOTAL_MONTO_FACTURADO - rec.TOTAL_MONTO_NC
			rec.ITBIS_POR_ADELANTAR = rec.ITBIS_TOTAL - rec.ITBIS_LLEVADO_ALCOSTO

			rec.pcount_01 = summary_dict["01"]["count"]
			rec.pcount_02 = summary_dict["02"]["count"]
			rec.pcount_03 = summary_dict["03"]["count"]
			rec.pcount_04 = summary_dict["04"]["count"]
			rec.pcount_05 = summary_dict["05"]["count"]
			rec.pcount_06 = summary_dict["06"]["count"]
			rec.pcount_07 = summary_dict["07"]["count"]
			rec.pcount_08 = summary_dict["08"]["count"]
			rec.pcount_09 = summary_dict["09"]["count"]
			rec.pcount_10 = summary_dict["10"]["count"]
			rec.pcount_11 = summary_dict["11"]["count"]

			rec.pamount_01 = summary_dict["01"]["amount"]
			rec.pamount_02 = summary_dict["02"]["amount"]
			rec.pamount_03 = summary_dict["03"]["amount"]
			rec.pamount_04 = summary_dict["04"]["amount"]
			rec.pamount_05 = summary_dict["05"]["amount"]
			rec.pamount_06 = summary_dict["06"]["amount"]
			rec.pamount_07 = summary_dict["07"]["amount"]
			rec.pamount_08 = summary_dict["08"]["amount"]
			rec.pamount_09 = summary_dict["09"]["amount"]
			rec.pamount_10 = summary_dict["10"]["amount"]
			rec.pamount_11 = summary_dict["11"]["amount"]

	@api.multi
	@api.depends("sale_report")
	def _sale_report_totals(self):

		# Tipos de NCFs by name
		summary_dict = {
			"final": {"count": 0, "amount": 0.0},
			"fiscal": {"count": 0, "amount": 0.0},
			"gov": {"count": 0, "amount": 0.0},
			"special": {"count": 0, "amount": 0.0},
			"unico": {"count": 0, "amount": 0.0},
		}

		for rec in self:

			if rec.name:
				report_month, report_year = rec.name.split("/")
			else:
				report_month = report_year = False

			rec.SALE_ITBIS_TOTAL = 0
			rec.SALE_ITBIS_NC = 0
			rec.SALE_ITBIS_CHARGED = 0
			rec.SALE_TOTAL_MONTO_FACTURADO = 0
			rec.SALE_TOTAL_MONTO_NC = 0
			rec.SALE_TOTAL_MONTO_CHARGED = 0
			rec.MONTO_FACTURADO_EXCENTO = 0

			for sale in rec.sale_report:

				ncf_year, ncf_month, ncf_day = sale.FECHA_COMPROBANTE.split("-")

				if sale.NUMERO_COMPROBANTE_FISCAL[9:-8] == "04":
					rec.SALE_ITBIS_NC += sale.ITBIS_FACTURADO
					rec.SALE_TOTAL_MONTO_NC += sale.MONTO_FACTURADO
					# TODO falta manejar las notas de credito que afectan facturas de otro periodo.
					rec.MONTO_FACTURADO_EXCENTO -= sale.MONTO_FACTURADO_EXCENTO
				else:

					if int(report_month) == int(
							ncf_month) and report_year == ncf_year:  # this validation is to avoid add amounts of invoices of previous months
						rec.SALE_TOTAL_MONTO_FACTURADO += sale.MONTO_FACTURADO
						rec.SALE_ITBIS_TOTAL += sale.ITBIS_FACTURADO
						rec.MONTO_FACTURADO_EXCENTO += sale.MONTO_FACTURADO_EXCENTO

				summary_dict[sale.invoice_id.sale_fiscal_type]["count"] += 1
				summary_dict[sale.invoice_id.sale_fiscal_type]["amount"] += sale.MONTO_FACTURADO

			rec.SALE_ITBIS_CHARGED = rec.SALE_ITBIS_TOTAL - rec.SALE_ITBIS_NC
			rec.SALE_TOTAL_MONTO_CHARGED = rec.SALE_TOTAL_MONTO_FACTURADO - rec.SALE_TOTAL_MONTO_NC

			rec.count_final = summary_dict["final"]["count"]

			rec.count_fiscal = summary_dict["fiscal"]["count"]
			rec.count_gov = summary_dict["gov"]["count"]
			rec.count_special = summary_dict["special"]["count"]
			rec.count_unico = summary_dict["unico"]["count"]
			rec.amount_final = summary_dict["final"]["amount"]
			rec.amount_fiscal = summary_dict["fiscal"]["amount"]
			rec.amount_gov = summary_dict["gov"]["amount"]
			rec.amount_special = summary_dict["special"]["amount"]
			rec.amount_unico = summary_dict["unico"]["amount"]

	@api.multi
	@api.depends("purchase_report", "sale_report")
	def _count_records(self):
		for rec in self:
			rec.COMPRAS_CANTIDAD_REGISTRO = rec.purchase_report and len(rec.purchase_report)
			rec.VENTAS_CANTIDAD_REGISTRO = rec.sale_report and len(rec.sale_report)
			rec.CANCEL_CANTIDAD_REGISTRO = rec.cancel_report and len(rec.cancel_report)
			rec.EXTERIOR_CANTIDAD_REGISTRO = rec.exterior_filename and len(rec.exterior_report)

	def get_invoice_in_draft_error(self, invoice_ids):
		error_list = {}
		error_msg = "Factura sin validar"
		for invoice_id in invoice_ids:
			if not error_list.get(invoice_id.id, False):
				error_list.update(
					{invoice_id.id: [
						(invoice_id.type, invoice_id.number, error_msg)]})
			else:
				error_list[invoice_id.id].append(
					(invoice_id.type, invoice_id.number, error_msg))
		return error_list

	''''
		With this method, we want get all invoices paid in a period of time (normally months later)
		and use them in the report of the current month (period: start date and end date given), but
		only if those invoices have retentions.
		Note for 606 report: acording with some accountants, this should be only valid for invoices
		with retention of ITBIS and ISR and of kind "Informal", which means that
		don't matter if the NCF is issued by a person or by a business,
		what matter is the document/identification of the provider,
		if this is of kind of "cedula", so it is informal.
	'''

	def get_late_paid_invoice_with_retentions(self, start_date, end_date):
		print('get_late_paid_invoice_with_retentions')
		invoices = self.env["account.invoice"]  # this is like define an empty array|object

		payments = self.env["account.payment"].search([
			('payment_date', '>=', start_date),
			('payment_date', '<=', end_date),
			('invoice_ids', '!=', False)  # TODO validate if this "where" is necessary.
		])

		for payment in payments:
			print('payment = ', payment)
			# RNC_CEDULA, TIPO_IDENTIFICACION = self.get_identification_info(payment.partner_id.vat)

			# if TIPO_IDENTIFICACION == "2": # just informal with or without ncf given.
			'''
				Here we want the latest payment, this means that this was the date
				when the invoice was paid fully.
			'''
			self.env.cr.execute(
				"select * from account_invoice_payment_rel where payment_id = %s order by payment_id desc limit 1" % payment.id)
			payment_rel = self.env.cr.dictfetchone()  # return just one diccionario, like laravel: ->first()
			# last_payment = self.env["account.payment"].browse(payment_rel['payment_id'])
			invoice = self.env["account.invoice"].browse(payment_rel['invoice_id'])

			# RETENCION_RENTA =  ITBIS_RETENIDO = False

			if invoice.type == 'in_invoice':  # 606
				print("invoice.type == 'in_invoice'")
				FECHA_PAGO, ITBIS_RETENIDO, RETENCION_RENTA, TIPO_RETENCION_ISR = self.get_payment_date_and_retention_data(
					invoice)

			if invoice.type == 'out_invoice':  # 607
				print("invoice.type == 'out_invoice'")
				FECHA_PAGO, ITBIS_RETENIDO = self.get_607_itbis_retenido_and_date(invoice)
				RETENCION_RENTA = False  # TODO need to be programmed for business and persons using "CÉDULA" as RNC, 'cause they can get ISR retentions

			if ITBIS_RETENIDO or RETENCION_RENTA:
				print("ITBIS_RETENIDO or RETENCION_RENTA")
				invoices |= invoice  # this is like array_push(), just making appends

		return invoices

	@api.model
	def get_type_identification(self, vat):
		rnc_schedule = vat and re.sub("[^0-9]", "", vat.strip()) or False
		print('get_type_identification ==> ', rnc_schedule)
		identification_type = "3"
		if rnc_schedule:
			if len(rnc_schedule) == 9:
				identification_type = "1"
			elif len(rnc_schedule) == 11:
				identification_type = "2"

		if identification_type == "3":
			rnc_schedule = ""

		return rnc_schedule, identification_type

	def validate_fiscal_information(self, vat, invoice):

		error_list = []

		if invoice.type == 'out_invoice' or invoice.number.startswith(
				'B11'):  # B11... are NCF issue by the own company, so validate them with company's RNC/CEDULA
			vat = invoice.company_id.vat

		if vat and len(vat) == 9 and not rnc.is_valid(vat):
			error_list.append(u"El RNC no es válido")

		if vat and len(vat) == 11 and not cedula.is_valid(vat):
			error_list.append(u"La Cédula no es válida")

		if not ncf.is_valid(invoice.number) or not ncf.check_dgii(vat, invoice.number):
			error_list.append(u"El NCF no es válido.  RNC: %s y tipo de Factura: %s" % (vat, invoice.type))

		if not invoice.refund_invoice_id and invoice.type in ("out_refund", "in_refund"):
			error_list.append(u"NC/ND sin comprobante que afecta")

		if not invoice.number:
			error_list.append(u"Factura validada sin número asignado")

		if invoice.type == 'in_invoice' and not invoice.expense_type:
			error_list.append(
				u"La factura %s no tiene especificado el tipo de costos y gastos requerído por el DGII." % invoice.number)

		return error_list

	# 608
	@api.multi
	def create_cancel_invoice_lines(self, cancel_invoice_ids):
		self.cancel_report.unlink()
		new_cancel_report = []
		cancel_line = 1
		for invoice_id in cancel_invoice_ids:
			new_cancel_report.append((0, 0, {"LINE": cancel_line, "TIPO_ANULACION": invoice_id.anulation_type,
											 "FECHA_COMPROBANTE": invoice_id.date_invoice,
											 "NUMERO_COMPROBANTE_FISCAL": invoice_id.move_name}))
			self.write({"cancel_report": new_cancel_report})
			cancel_line += 1

	@api.multi
	def get_numero_de_comprobante_modificado(self, invoice_id):

		NUMERO_COMPROBANTE_MODIFICADO = False
		AFFECTED_NVOICE_ID = False

		origin_invoice_id = invoice_id.refund_invoice_id.filtered(lambda x: x.state in ("open", "paid"))

		if not origin_invoice_id:
			origin_invoice_id = self.env["account.invoice"].search([('number', '=', invoice_id.origin)])

		NUMERO_COMPROBANTE_MODIFICADO = origin_invoice_id[0].number
		AFFECTED_NVOICE_ID = origin_invoice_id[0].id

		return NUMERO_COMPROBANTE_MODIFICADO, AFFECTED_NVOICE_ID

	'''
		This method is only called when the Invoice is paid.
		This method is used for 606 report.  There is other
		method similar to this but for 607 report, its name is:
		get_607_itbis_retenido_and_date().  We decide to keep these
		two method separate for better understand and convenience.
	'''

	def get_payment_date_and_retention_data(self, invoice):

		FECHA_PAGO = False
		ITBIS_RETENIDO = 0
		RETENCION_RENTA = 0
		TIPO_RETENCION_ISR = False

		# payment_rel = self.env["account.invoice.payment.rel"].search(['invoice', '=', invoice.id]) # This return an error:  KeyError: 'account.invoice.payment.rel'
		self.env.cr.execute("select * from account_invoice_payment_rel where invoice_id = %s" % invoice.id)
		payment_rel = self.env.cr.dictfetchone()  # return just one diccionario, like laravel: ->first()

		if invoice.number.startswith('B04'):  # This is a Credit Note
			'''
			#TODO validate with an accountant if Credit Note require payment date.
			# If true so this is the same date when the NC was made.
			By now, one accoutant (Henry) said that he think could be the same date as NC or could be leave empty. (Aug 14th, 2018)
			'''
			FECHA_PAGO = invoice.date_invoice

		elif payment_rel:

			payment = self.env["account.payment"].browse(payment_rel['payment_id'])
			FECHA_PAGO = payment.payment_date

		else:  # might be a paid with a "NOTA DE CREDITO"
			'''
			#TODO este else quizás no debería ser alcanzado dado que una factura no se puede pagar con una NC, en teoría...
			pues no te darán una NC de una factura que no está pagada y por lo consiguiente si una factura fue pagada debe tener su forma de pago
			que NO es una nota de crédito.   Quizás la opción de pago 06 = NOTA DE CREDITO del 606 es para ponerle a las NC como tal.
			Update 1:  en Aug 14th, 2018 el contable Henry dice que si es posible esto dado que la factura puede ser a crédito de 30 o 90 días y luego el cliente
			le pide al proveedor que le reembolse parte de esa factura por algún error.
			'''

			refund_invoice_id = self.env["account.invoice"].search([('refund_invoice_id', '=', invoice.id)], limit=1,
																   order='refund_invoice_id desc')  # the last one is the real payment day

			if refund_invoice_id:  # this is the Credit Notes
				FECHA_PAGO = refund_invoice_id.date_invoice

		move_id = self.env["account.move.line"].search(
			[("move_id", "=", invoice.move_id.id), ('full_reconcile_id', '!=', False)])  # just one is full_reconcile_id

		if invoice.journal_id.purchase_type in ("informal", "normal"):

			if move_id:

				account_move_lines = self.env["account.move.line"].search([
					('move_id', '=', invoice.move_id.id),
					('tax_line_id', '!=', False)
				])  # I removed the filter ('payment_id', '!=', False) because in one of my case the move lines don't have payment_id, why?, I don't have idea....

				if account_move_lines:
					for line in account_move_lines:
						if line.tax_line_id.purchase_tax_type == "ritbis":
							ITBIS_RETENIDO += line.credit
						elif line.tax_line_id.purchase_tax_type == "isr":
							RETENCION_RENTA += line.credit
							TIPO_RETENCION_ISR = line.tax_line_id.isr_retention_type or None

		return FECHA_PAGO, ITBIS_RETENIDO, RETENCION_RENTA, TIPO_RETENCION_ISR

	'''
		This method return:
		Impuesto Selectivo al Consumo (Casilla 20),
		Otros Impuesto/Tasas (Casilla 21),
		and Monto Propina Legal (Casilla 22)
		but only when the invoice is Open or Paid
	'''

	def get_isc_propina_otros(self, invoice_id):

		IMPUESTO_ISC = 0
		IMPUESTOS_OTROS = 0
		MONTO_PROPINA_LEGAL = 0

		if invoice_id.id == False:  # TODO for some reason, invoice_id has not any properties some times...
			return IMPUESTO_ISC, IMPUESTOS_OTROS, MONTO_PROPINA_LEGAL

		if invoice_id.state in ("open", "paid"):

			account_move_lines = self.env["account.move.line"].search(
				[('move_id', '=', invoice_id.move_id.id), ('tax_line_id', '!=', False)])

			if account_move_lines:
				for line in account_move_lines:
					if line.tax_line_id.purchase_tax_type == "isc":
						IMPUESTO_ISC += line.debit  # TODO ask to accountant if the field should be debit or credit, by now I am seeting those value in debit field
					elif line.tax_line_id.purchase_tax_type in (
							"cdt"):  # TODO might be there another taxes as "IMPUESTOS_OTROS" that are not just CDT.
						IMPUESTOS_OTROS += line.debit  # TODO ask to accountant if the field should be debit or credit, by now I am seeting those value in debit field
					elif line.tax_line_id.purchase_tax_type in ("propina_legal"):
						MONTO_PROPINA_LEGAL += line.debit  # TODO ask to accountant if the field should be debit or credit, by now I am seeting those value in debit field

		return IMPUESTO_ISC, IMPUESTOS_OTROS, MONTO_PROPINA_LEGAL

	def get_payment_form_of_purchase(self, invoice):

		FORMA_PAGO = '04'  # 04 = COMPRA A CREDITO

		if invoice.state == "paid":

			self.env.cr.execute("select * from account_invoice_payment_rel where invoice_id = %s" % invoice.id)
			payment_rel = self.env.cr.dictfetchall()  # return an array of dicts, like laravel: ->get()

			if invoice.number.startswith('B04') or invoice.number[9:11] == '04':  # This is a Credit Note
				'''
				#TODO validate with an accountant if Credit Note require Payment Method.
				By now, one accoutant (Henry) said that he think could be the same payment method as original invoice or could be leave empty. (Aug 14th, 2018).
				But, I think it need be just Credit Note 'cause you don't use Cash or Credit Card to pay a NC (Manuel González)
				Update 1: in Aug 28, a DGII's employee (one those that work in "fiscalización") says that it need to be the original invoice payment method,
				but that employee seems not be very sure about it.  But due two "confirmations", I am going to set the original invoice payment method
				'''
				FORMA_PAGO = self.get_payment_form_of_purchase(invoice.refund_invoice_id)

			elif not payment_rel:  # could be a NOTA DE CREDITO, they don't seems store payment_id
				'''
				#TODO este else quizás no debería ser alcanzado dado que una factura no se puede pagar con una NC, en teoría...
				pues no te darán una NC de una factura que no está pagada y por lo consiguiente si una factura fue pagada debe tener su forma de pago
				que NO es una nota de crédito.   Quizás la opción de pago 06 = NOTA DE CREDITO del 606 es para ponerle a las NC como tal.
				Update 1:  en Aug 14th, 2018 el contable Henry dice que si es posible esto dado que la factura puede ser a crédito de 30 o 90 días y luego el cliente
				le pide al proveedor que le reembolse parte de esa factura por algún error.
				Notar:
				- Si este elif es alcanzado es porque la factura fue pagada full con NC.
				- Al ser pagada con NC y estar en este punto (sin pago registrado) signfica que el bien/servicio recibido fue a crédito
				- Si la NC es parcial (no es el valor total de está factura), se alcanzará el else más abajo y la forma de pago podría ser una de las opciones de las validaciones hechas allí.
				Att: Manuel González <manuel@softnet.do>
				'''

				# refund_invoice_id = self.env["account.invoice"].search([('refund_invoice_id', '=', invoice.id)])
				# if refund_invoice_id:
				#     FORMA_PAGO = '06' # 06 = NOTA DE CREDITO

				FORMA_PAGO = '04'  # 04 = COMPRA A CREDITO

			elif len(payment_rel) > 1:

				FORMA_PAGO = '07'  # 07 = MIXTO

			else:

				payment = self.env["account.payment"].browse(payment_rel[0]['payment_id'])

				if payment.writeoff_account_id.id > 0:  # TODO validate with an accountant this (lo que se debe validar es si una factura es de mil pesos pagan solo 500 y se hace un writeoff de los otros 500...).
					FORMA_PAGO = '07'  # 07 = MIXTO
				elif payment.journal_id.payment_form == 'cash':
					FORMA_PAGO = '01'
				elif payment.journal_id.payment_form == 'bank':
					FORMA_PAGO = '02'
				elif payment.journal_id.payment_form == 'card':
					FORMA_PAGO = '03'
				elif payment.journal_id.payment_form == 'credit':  # just in case they have a journal of credit
					FORMA_PAGO = '04'
				elif payment.journal_id.payment_form == 'swap':
					FORMA_PAGO = '05'  # Permuta

		return FORMA_PAGO

	@api.multi
	def create_sales_lines(self, data):
		"""
		Function create sales line report 607.
		:param data:
		:return:
		"""
		for d in data:
			self.env['dgii.report.sale.line'].create(d)

	@api.multi
	def create_purchase_lines(self, data):
		"""
		Function create purchase line report 606.
		:param data:
		:return:
		"""
		for d in data:
			self.env['dgii.report.purchase.line'].create(d)

	@api.multi
	def post_error_list(self, error_list):
		out_inovice_url = "/web#id={}&view_type=form&model=account.invoice&action=196"
		in_inovice_url = "/web#id={}&view_type=form&model=account.invoice&menu_id=119&action=197"
		if error_list:
			message = "<ul>"
			for ncf, errors in error_list.iteritems():
				message += "<li>{}</li><ul>".format(errors[0][1] or "Factura invalida")
				for error in errors:
					if error[0] in ("out_invoice", "out_refund"):
						message += u"<li><a target='_blank' href='{}'>{}</a></li>".format(out_inovice_url.format(ncf),
																						  error[2])
					else:
						message += u"<li><a target='_blank' href='{}'>{}</a></li>".format(in_inovice_url.format(ncf),
																						  error[2])
				message += "</ul>"
			message += "</ul>"

			self.message_post(body=message)
			self.state = "error"
		else:
			self.message_post(body="Generado correctamente")
			self.state = "done"

	'''
		Only call this method when the invoice is paid.
	'''

	def get_607_itbis_retenido_and_date(self, invoice):

		FECHA_RETENCION = None
		ITBIS_RETENIDO_POR_TERCEROS = None

		'''
		#TODO the below query return the last payment to the invoice.
		Whether an invoice has multiple payments, the most natural is that the last payment is that one with
		writeoff_account_id property, because if a customer is making multiple payment to you, the most natural is
		that you don't register any tax retention until the invoice is full paid.  So take care about this.
		Att: Manuel Gonzalez <manuel@softnet.do> Ago 25, 2018.
		'''
		self.env.cr.execute(
			"select * from account_invoice_payment_rel where invoice_id = %s order by payment_id desc limit 1" % invoice.id)
		payment_rel = self.env.cr.dictfetchone()  # return just one diccionario, like laravel: ->first()

		if payment_rel:

			payment = self.env["account.payment"].browse(payment_rel['payment_id'])

			if payment.writeoff_account_id:  # this payment could have retentions...
				'''
					By default the account ID 100 is "ITBIS Retenido Persona Jurídica (N 02-05)"
					and for a company with RNC, normally this is the kind of retentions that they have...
					#TODO need be programed and tested with "Proveedores Informales" giving NCF
					But you know, some accountant could change the account and this default ID could be other
					and for this reason we set a new field in account_account model with name sale_tax_type
					and wih this avoid any confusion.
					There were other way to filter this without a new field in account_account model, but that way
					is so confused and the new field's solution is more direct and clear.
				'''

				if payment.writeoff_account_id.sale_tax_type == 'ritbis_pjuridica_n_02_05':
					'''
						So, go ahead and look for the retention amount in move lines...
						In the below query, we don't search by invoice_id because normally there are just 3 rows asociated with an invoice
						and none of them have the account_id that we need to filter (in ODOO 10), instead we are searching
						by the field "ref" because it is the only way that we can do it.  And we use invoice.move_name to filter because
						we think this is the correct way (#TODO validate this...) although normally the invoice.move_name = invoice.number
					'''

					account_move_line = self.env["account.move.line"].search(
						[('ref', '=', invoice.move_name), ('account_id', '=', payment.writeoff_account_id.id)])

					if account_move_line:
						FECHA_RETENCION = payment.payment_date  # in practical terms, this is  "FECHA DE RETENCIÓN" in 607 report.
						ITBIS_RETENIDO_POR_TERCEROS = account_move_line.debit  # TODO - We wait just one record, but take care, maybe could be more than one in some use cases what was no tested.

		return FECHA_RETENCION, ITBIS_RETENIDO_POR_TERCEROS

	'''
		Call this method only when the invoice is paid.
	'''

	def get_forma_pago_ventas(self, invoice, commun_data):

		self.env.cr.execute("select * from account_invoice_payment_rel where invoice_id = %s" % invoice.id)
		payment_rel = self.env.cr.dictfetchall()  # return an array of dicts, like laravel: ->get()

		if invoice.number.startswith('B04') or (invoice.number[9:11] == '04' and len(
				invoice.number) > 11):  # This is a Credit Note, the len validation if to avoid false positive with invoice like 'B0100000004'
			'''
			#TODO validate with an accountant if Credit Note require "Payment Method".
			By now, one accoutant (Henry) said that by logic, NC should have the same payment method as original invoice,
			but he also says that in a 607 report, he sent a NC with "Payment Method" as empty and all was OK.
			So, by now just let it empty, is more "logic" to me. (Manuel González <manuel@softnet.do>)
			'''
			commun_data['MONTOS_A_CREDITO'] = 0  # set 0 'cause by default the invoice comes with this value

		elif not payment_rel:  # could be full paid with a "NOTA DE CREDITO", they don't seems store payment_id
			'''
			#TODO este else quizás no debería ser alcanzado dado que una factura no se puede pagar con una NC, en teoría...
			pues no deberíamos dar una NC de una factura que no está pagada y por lo consiguiente si una factura fue pagada debe tener su forma de pago
			que NO es una nota de crédito.
			Update 1:  en Aug 14th, 2018 el contable Henry dice que si es posible esto dado que la factura puede ser a crédito de 30 o 90 días y luego el cliente
			le pide al proveedor que le reembolse parte de esa factura por algún error.  Comentario de Manuel Gonzalez <manuel@softnet.do>: Aunque de todas formas
			si este caso se da, solo se alcanzaría este elif cuando el pago con la nota de crédito sea por el monto total de la factura, si el la NC fue parcial
			significa que alguna otra forma de pago fue usada y entonces se iría al else de abajo.
			Entonces, la única "forma de pago" factible aquí sería "A Crédito" pues la factura nunca fue pagada realmente.  Aunque si este elif es alcanzado significa
			que la NC emitida para esta factura fue hecha en el mismo período que está factura; si la NC se emite en un período posterior, pues está factura no irá, solo
			aparecera como "NÚMERO DE COMPROBANTE MODIFICADO".  Att: Manuel González <manuel@softnet.do>
			'''

			commun_data['MONTOS_A_CREDITO'] = 0  # set 0 'cause by default the invoice comes with this value

		else:

			for prel in payment_rel:

				payment = self.env["account.payment"].browse(prel['payment_id'])

				if payment.journal_id.payment_form == 'cash':
					commun_data['MONTOS_PAGADOS_EFECTIVO'] += payment.amount
				elif payment.journal_id.payment_form == 'bank':
					commun_data['MONTOS_PAGADOS_BANCO'] += payment.amount
				elif payment.journal_id.payment_form == 'card':
					commun_data['MONTOS_PAGADOS_TARJETAS'] += payment.amount
				elif payment.journal_id.payment_form == 'credit':  # just in case they have a journal of credit
					commun_data['MONTOS_A_CREDITO'] += payment.amount
				elif payment.journal_id.payment_form == 'bond':
					commun_data['MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS'] += payment.amount
				elif payment.journal_id.payment_form == 'swap':
					commun_data['MONTOS_EN_PERMUTA'] += payment.amount
				else:
					commun_data['MONTOS_EN_OTRAS_FORMAS_VENTAS'] += payment.amount  # like Bitcoin and others

			'''
				This is not going to 607 report or any model,
				just use to do operations
			'''
			commun_data['GRAN_TOTAL_PAGADO'] = commun_data['MONTOS_PAGADOS_EFECTIVO'] \
											   + commun_data['MONTOS_PAGADOS_BANCO'] + commun_data[
												   'MONTOS_PAGADOS_TARJETAS'] \
											   + commun_data['MONTOS_A_CREDITO'] + commun_data[
												   'MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS'] \
											   + commun_data['MONTOS_EN_PERMUTA'] + commun_data[
												   'MONTOS_EN_OTRAS_FORMAS_VENTAS']

		return commun_data

	def get_607_report_data(self, invoice, commun_data):

		commun_data['TIPO_DE_INGRESO'] = invoice.income_type
		commun_data[
			'MONTOS_A_CREDITO'] = invoice.amount_total_signed  # by default it is credit.  #TODO, there are too: amount_total_company_signed and amount_total. What are the differences?

		if invoice.state == "paid":

			FECHA_RETENCION, ITBIS_RETENIDO_POR_TERCEROS = self.get_607_itbis_retenido_and_date(invoice)
			formas_pagos = self.get_forma_pago_ventas(invoice, commun_data)

			commun_data = dict(commun_data,
							   **formas_pagos)  # with this, we merge two dict.  All keys's values are overritten from A (commun_data) to what is set on B (formas_pagos)

			commun_data['FECHA_RETENCION'] = FECHA_RETENCION
			commun_data['ITBIS_RETENIDO_POR_TERCEROS'] = ITBIS_RETENIDO_POR_TERCEROS
			commun_data[
				'MONTOS_A_CREDITO'] = 0  # if the invoice is paid, start with 0, if there is remaining amount, then below it is calculated and assigned.

			'''
				Avoid set "ITBIS RETENIDO POR TERCEROS" and any payment form instead
				"A CRÉDITO" for invoice paid months laters (and other stuff).

				This case happen when for example an invoice was issue on June 2018,
				then the customer paid it on July 2018 and he made retentions.
				If you come back to June reports and re-generate it (or if ODOO re-generate it when you enter to it)
				Then, without this validation below, you should see in the report for June those retentions,
				what is wrong because the invoice wasn't paid in that month.
			'''
			month, year = self.name.split("/")
			invoiceMonth = int(invoice.date_invoice[5:7])
			retentionMonth = int(FECHA_RETENCION[5:7]) if FECHA_RETENCION else False
			periodMonth = int(month)

			if retentionMonth and invoiceMonth != retentionMonth and invoiceMonth == periodMonth:
				commun_data['FECHA_RETENCION'] = None
				commun_data['ITBIS_RETENIDO_POR_TERCEROS'] = 0
				commun_data['MONTOS_A_CREDITO'] = invoice.amount_total_signed
				commun_data['MONTOS_PAGADOS_EFECTIVO'] = commun_data['MONTOS_PAGADOS_BANCO'] = commun_data[
					'MONTOS_PAGADOS_TARJETAS'] = commun_data['MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS'] = commun_data[
					'MONTOS_EN_PERMUTA'] = commun_data['MONTOS_EN_OTRAS_FORMAS_VENTAS'] = 0

			gran_total_paid_plus_retentions = commun_data['GRAN_TOTAL_PAGADO'] + commun_data[
				'ITBIS_RETENIDO_POR_TERCEROS'] if commun_data['ITBIS_RETENIDO_POR_TERCEROS'] else commun_data[
				'GRAN_TOTAL_PAGADO']

			if invoice.amount_total_signed > gran_total_paid_plus_retentions:
				'''
					The accountant Henry says that an invoice could be partially paid with any valid
					Payment Method and the remaining amount could be as "A Crédito"
					#TODO here also need be calculate the column 12 607 report "Retención Renta por Terceros"
					that by now we set that column as 0, but if this report is going to be used by someone with
					CÉDULA instead of RNC or if its business is one of those that have "Retención Renta por Terceros",
					so that is something to be programmed.
				'''
				commun_data['MONTOS_A_CREDITO'] = invoice.amount_total_signed - gran_total_paid_plus_retentions

		return commun_data

	@api.multi
	def generate_report(self):
		print("generate_report")
		try:
			month, year = self.name.split("/")
			last_day = calendar.monthrange(int(year), int(month))[1]
			start_date = "{}-{}-01".format(year, month)
			end_date = "{}-{}-{}".format(year, month, last_day)
		except Exception:
			raise exceptions.ValidationError(u"Período inválido")

		self.purchase_report.unlink()
		self.sale_report.unlink()
		self.cancel_report.unlink()
		self.exterior_report.unlink()

		self.it_filename = False
		self.it_binary = False
		self.ir17_filename = False
		self.ir17_binary = False

		self.sale_binary = False
		self.sale_filename = False
		self.purchase_binary = False
		self.purchase_filename = False
		self.cancel_binary = False
		self.cancel_filename = False

		xls_dict = {"it1": {}, "ir17": {}}
		purchase_report = []
		sale_report = []
		ext_report = []
		sale_line = 1
		purchase_line = 1
		ext_line = 1

		sale_except_tax_id = self.env.ref("l10n_do.{}_tax_0_sale".format(self.company_id.id))
		purchase_except_tax_id = self.env.ref("l10n_do.{}_tax_0_purch".format(self.company_id.id))
		untax_ids = (sale_except_tax_id.id, purchase_except_tax_id.id)

		journal_ids = self.env["account.journal"].search(
			['|', ('ncf_control', '=', True), ('ncf_remote_validation', '=', True)])
		print('journal_ids = ', journal_ids, 'count = ', len(journal_ids))

		# searching invoices to this period
		invoice_ids = self.env["account.invoice"].search([
			('date_invoice', '>=', start_date),
			('date_invoice', '<=', end_date),
			('journal_id', 'in', journal_ids.ids),
		], order='date_invoice, number asc')
		print('invoice_ids 1 = ', invoice_ids, 'count = ', len(invoice_ids))

		error_list = self.get_invoice_in_draft_error(invoice_ids.filtered(lambda x: x.state == "draft"))

		self.create_cancel_invoice_lines(invoice_ids.filtered(lambda x: x.state == 'cancel' and
																		x.type in ("out_invoice", "out_refund") and
																		x.move_name))

		invoice_ids = invoice_ids.filtered(lambda x: x.state in ('open', 'paid'))
		print('invoice_ids 2 = ', invoice_ids, 'count = ', len(invoice_ids))
		invoice_ids |= self.get_late_paid_invoice_with_retentions(start_date, end_date)
		print('invoice_ids 3 = ', invoice_ids, 'count = ', len(invoice_ids))
		count = len(invoice_ids)

		# *****************************  START FOR EACH INVOICE *****************************
		for invoice_id in invoice_ids:
			print("invoice_id = ", invoice_id)
			print("invoice_id.state = ", invoice_id.state)
			vat = self.get_type_identification(invoice_id.partner_id.vat)

		# *****************************  END FOR EACH INVOICE *******************************



	def generate_xls_files(self, xls_dict):
		# fill IT-1 excel file
		cwf = os.path.join(os.path.dirname(os.path.abspath(__file__)), "IT-1-2017.xlsx")
		wb = load_workbook(cwf)
		ws1 = wb["IT-1"]  # Get sheet 1 in writeable copy
		xls_dict["it1"].update({"S43": self.positive_balance})
		for k, v in xls_dict["it1"].iteritems():
			ws1[k] = v

		period = self.name.split("/")
		FILENAME = "IT-1-{}-{}.xlsx".format(period[0], period[1])
		wb.save("/tmp/{}".format(FILENAME))
		with open("/tmp/{}".format(FILENAME), "rb") as xls_file:
			self.write({
				'it_filename': FILENAME,
				'it_binary': base64.b64encode(xls_file.read())
			})

		# fill IR-17 excel file
		cwf = os.path.join(os.path.dirname(os.path.abspath(__file__)), "IR-17-2015.xlsx")
		wb = load_workbook(cwf)
		ws1 = wb["IR17"]  # Get sheet 1 in writeable copy
		for k, v in xls_dict["ir17"].iteritems():
			ws1[k] = v

		period = self.name.split("/")
		FILENAME = "IR-17-{}-{}.xlsx".format(period[0], period[1])
		wb.save("/tmp/{}".format(FILENAME))
		with open("/tmp/{}".format(FILENAME), "rb") as xls_file:
			self.write({
				'ir17_filename': FILENAME,
				'ir17_binary': base64.b64encode(xls_file.read())
			})

	def generate_txt_files_norma_06_18(self):

		company_fiscal_identificacion = re.sub("[^0-9]", "", self.company_id.vat)
		period = self.name.split("/")
		month = period[0]
		year = period[1]

		''' ************************ 607 TXT REPORT ******************************** '''
		sale_path = '/tmp/607{}.txt'.format(company_fiscal_identificacion)
		sale_file = open(sale_path, 'w')

		lines = []

		CANTIDAD_REGISTRO = len(self.sale_report)

		header = "607"
		header += "|"
		header += company_fiscal_identificacion + "|"
		header += str(year)
		header += str(month).zfill(2)
		header += "|"
		header += str(CANTIDAD_REGISTRO)

		lines.append(header)

		for sale_line in self.sale_report:
			ln = ""
			ln += sale_line.RNC_CEDULA + "|"
			ln += sale_line.TIPO_IDENTIFICACION + "|"
			ln += sale_line.NUMERO_COMPROBANTE_FISCAL + "|"
			ln += sale_line.NUMERO_COMPROBANTE_MODIFICADO + "|" if sale_line.NUMERO_COMPROBANTE_MODIFICADO else "|"
			ln += str(sale_line.TIPO_DE_INGRESO) + "|"
			ln += sale_line.FECHA_COMPROBANTE.replace("-", "") + "|"
			ln += sale_line.FECHA_RETENCION.replace("-", "") + "|" if sale_line.FECHA_RETENCION else "|"
			ln += str(abs(sale_line.MONTO_FACTURADO)) + "|"
			ln += str(abs(sale_line.ITBIS_FACTURADO)) + "|" if sale_line.ITBIS_FACTURADO else "|"
			ln += str(
				abs(sale_line.ITBIS_RETENIDO_POR_TERCEROS)) + "|" if sale_line.ITBIS_RETENIDO_POR_TERCEROS else "|"
			ln += str(abs(sale_line.ITBIS_PERCIBIDO)) + "|" if sale_line.ITBIS_PERCIBIDO else "|"
			ln += str(
				abs(sale_line.RETENCION_RENTA_POR_TERCEROS)) + "|" if sale_line.RETENCION_RENTA_POR_TERCEROS else "|"
			ln += str(abs(sale_line.ISR_PERCIBIDO)) + "|" if sale_line.ISR_PERCIBIDO else "|"
			ln += str(abs(sale_line.IMPUESTO_ISC)) + "|" if sale_line.IMPUESTO_ISC else "|"
			ln += str(abs(sale_line.IMPUESTOS_OTROS)) + "|" if sale_line.IMPUESTOS_OTROS else "|"
			ln += str(abs(sale_line.MONTO_PROPINA_LEGAL)) + "|" if sale_line.MONTO_PROPINA_LEGAL else "|"
			ln += str(abs(sale_line.MONTOS_PAGADOS_EFECTIVO)) + "|" if sale_line.MONTOS_PAGADOS_EFECTIVO else "|"
			ln += str(abs(sale_line.MONTOS_PAGADOS_BANCO)) + "|" if sale_line.MONTOS_PAGADOS_BANCO else "|"
			ln += str(abs(sale_line.MONTOS_PAGADOS_TARJETAS)) + "|" if sale_line.MONTOS_PAGADOS_TARJETAS else "|"
			ln += str(abs(sale_line.MONTOS_A_CREDITO)) + "|" if sale_line.MONTOS_A_CREDITO else "|"
			ln += str(abs(
				sale_line.MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS)) + "|" if sale_line.MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS else "|"
			ln += str(abs(sale_line.MONTOS_EN_PERMUTA)) + "|" if sale_line.MONTOS_EN_PERMUTA else "|"
			ln += str(abs(sale_line.MONTOS_EN_OTRAS_FORMAS_VENTAS)) if sale_line.MONTOS_EN_OTRAS_FORMAS_VENTAS else ""
			lines.append(ln)

		for line in lines:
			sale_file.write(line + "\n")

		sale_file.close()
		sale_file = open(sale_path, 'rb')
		sale_binary = base64.b64encode(sale_file.read())
		report_name = 'DGII_F_607_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
		self.write({'sale_binary': sale_binary, 'sale_filename': report_name})

		''' ************************ 606 TXT REPORT ******************************** '''
		pruchase_path = '/tmp/606{}.txt'.format(company_fiscal_identificacion)
		purchase_file = open(pruchase_path, 'w')
		lines = []

		CANTIDAD_REGISTRO = len(self.purchase_report)

		header = "606"
		header += "|"
		header += company_fiscal_identificacion + "|"
		header += str(year)
		header += str(month).zfill(2)
		header += "|"
		header += str(CANTIDAD_REGISTRO)

		lines.append(header)

		for line in self.purchase_report:
			ln = ""
			ln += line.RNC_CEDULA + "|"
			ln += line.TIPO_IDENTIFICACION + "|"
			ln += line.TIPO_BIENES_SERVICIOS_COMPRADOS + "|"
			ln += line.NUMERO_COMPROBANTE_FISCAL + "|"
			ln += line.NUMERO_COMPROBANTE_MODIFICADO + "|" if line.NUMERO_COMPROBANTE_MODIFICADO else "|"
			ln += line.FECHA_COMPROBANTE.replace("-", "") + "|"
			ln += line.FECHA_PAGO.replace("-", "") + "|" if line.FECHA_PAGO else "" + "|"
			ln += str(abs(line.MONTO_FACTURADO_SERVICIOS)) + "|" if line.MONTO_FACTURADO_SERVICIOS else "|"
			ln += str(abs(line.MONTO_FACTURADO_BIENES)) + "|" if line.MONTO_FACTURADO_BIENES else "|"
			ln += str(abs(line.MONTO_FACTURADO)) + "|"  # the total
			ln += str(abs(line.ITBIS_FACTURADO_TOTAL)) + "|" if line.ITBIS_FACTURADO_TOTAL else "0" + "|"
			ln += str(abs(line.ITBIS_RETENIDO)) + "|" if line.ITBIS_RETENIDO else "|"
			ln += str(abs(line.ITBIS_SUJETO_PROPORCIONALIDAD)) + "|" if line.ITBIS_SUJETO_PROPORCIONALIDAD else "|"
			ln += str(abs(line.ITBIS_LLEVADO_ALCOSTO)) + "|" if line.ITBIS_LLEVADO_ALCOSTO else "|"
			ln += str(abs(line.ITBIS_POR_ADELANTAR)) + "|" if line.ITBIS_POR_ADELANTAR else "0" + "|"
			ln += str(abs(line.ITBIS_PERCIBIDO_COMPRAS)) + "|" if line.ITBIS_PERCIBIDO_COMPRAS else "|"
			ln += line.TIPO_RETENCION_ISR + "|" if line.TIPO_RETENCION_ISR else "|"
			ln += str(abs(line.RETENCION_RENTA)) + "|" if line.RETENCION_RENTA else "|"
			ln += str(abs(line.ISR_PERCIBIDO_COMPRAS)) + "|" if line.ISR_PERCIBIDO_COMPRAS else "|"
			ln += str(abs(line.IMPUESTO_ISC)) + "|" if line.IMPUESTO_ISC else "|"
			ln += str(abs(line.IMPUESTOS_OTROS)) + "|" if line.IMPUESTOS_OTROS else "|"
			ln += str(abs(line.MONTO_PROPINA_LEGAL)) + "|" if line.MONTO_PROPINA_LEGAL else "|"
			ln += line.FORMA_PAGO
			lines.append(ln)

		for line in lines:
			purchase_file.write(line + "\n")

		purchase_file.close()
		purchase_file = open(pruchase_path, 'rb')
		purchase_binary = base64.b64encode(purchase_file.read())
		purchase_filename = 'DGII_F_606_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year),
															str(month).zfill(2))
		self.write({'purchase_binary': purchase_binary, 'purchase_filename': purchase_filename})

		''' ************************ 608 TXT REPORT ******************************** '''
		path = '/tmp/608{}.txt'.format(company_fiscal_identificacion)
		file = open(path, 'w')
		lines = []

		header = "608"
		header += company_fiscal_identificacion.zfill(11)
		header += str(year)
		header += str(month).zfill(2)
		lines.append(header)

		for line in self.cancel_report:
			ln = ""
			ln += line.NUMERO_COMPROBANTE_FISCAL
			ln += line.FECHA_COMPROBANTE and line.FECHA_COMPROBANTE.replace("-", "") or ""
			ln += "{}".format(line.TIPO_ANULACION).zfill(2)
			lines.append(ln)

		for line in lines:
			file.write(line + "\n")

		file.close()
		file = open(path, 'rb')
		report = base64.b64encode(file.read())
		report_name = 'DGII_608_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
		self.write({'cancel_binary': report, 'cancel_filename': report_name})

	def generate_txt_files_norma_06_18_v2(self):

		company_fiscal_identificacion = re.sub("[^0-9]", "", self.company_id.vat)
		period = self.name.split("/")
		month = period[0]
		year = period[1]

		''' ************************ 607 TXT REPORT ******************************** '''
		sale_path = '/tmp/607{}.txt'.format(company_fiscal_identificacion)
		sale_file = open(sale_path, 'w')

		pipe = "|"
		lines = []

		CANTIDAD_REGISTRO = len(self.sale_report)

		header = "607"
		header += "|"
		header += company_fiscal_identificacion + "|"
		header += str(year)
		header += str(month).zfill(2)
		header += "|"
		header += str(CANTIDAD_REGISTRO)

		lines.append(header)

		for sale_line in self.sale_report:
			ln = ""
			ln += sale_line.RNC_CEDULA + "|"
			ln += sale_line.TIPO_IDENTIFICACION + "|"
			ln += sale_line.NUMERO_COMPROBANTE_FISCAL + "|"
			ln += sale_line.NUMERO_COMPROBANTE_MODIFICADO + "|" if sale_line.NUMERO_COMPROBANTE_MODIFICADO else "|"
			ln += str(sale_line.TIPO_DE_INGRESO) + "|"
			ln += sale_line.FECHA_COMPROBANTE.replace("-", "") + "|"
			ln += sale_line.FECHA_RETENCION.replace("-", "") + "|" if sale_line.FECHA_RETENCION else "|"
			ln += str(abs(sale_line.MONTO_FACTURADO)) + "|"
			ln += str(abs(sale_line.ITBIS_FACTURADO)) + "|" if sale_line.ITBIS_FACTURADO else "|"
			ln += str(
				abs(sale_line.ITBIS_RETENIDO_POR_TERCEROS)) + "|" if sale_line.ITBIS_RETENIDO_POR_TERCEROS else "|"
			ln += str(abs(sale_line.ITBIS_PERCIBIDO)) + "|" if sale_line.ITBIS_PERCIBIDO else "|"
			ln += str(
				abs(sale_line.RETENCION_RENTA_POR_TERCEROS)) + "|" if sale_line.RETENCION_RENTA_POR_TERCEROS else "|"
			ln += str(abs(sale_line.ISR_PERCIBIDO)) + "|" if sale_line.ISR_PERCIBIDO else "|"
			ln += str(abs(sale_line.IMPUESTO_ISC)) + "|" if sale_line.IMPUESTO_ISC else "|"
			ln += str(abs(sale_line.IMPUESTOS_OTROS)) + "|" if sale_line.IMPUESTOS_OTROS else "|"
			ln += str(abs(sale_line.MONTO_PROPINA_LEGAL)) + "|" if sale_line.MONTO_PROPINA_LEGAL else "|"
			ln += str(abs(sale_line.MONTOS_PAGADOS_EFECTIVO)) + "|" if sale_line.MONTOS_PAGADOS_EFECTIVO else "|"
			ln += str(abs(sale_line.MONTOS_PAGADOS_BANCO)) + "|" if sale_line.MONTOS_PAGADOS_BANCO else "|"
			ln += str(abs(sale_line.MONTOS_PAGADOS_TARJETAS)) + "|" if sale_line.MONTOS_PAGADOS_TARJETAS else "|"
			ln += str(abs(sale_line.MONTOS_A_CREDITO)) + "|" if sale_line.MONTOS_A_CREDITO else "|"
			ln += str(abs(
				sale_line.MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS)) + "|" if sale_line.MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS else "|"
			ln += str(abs(sale_line.MONTOS_EN_PERMUTA)) + "|" if sale_line.MONTOS_EN_PERMUTA else "|"
			ln += str(abs(sale_line.MONTOS_EN_OTRAS_FORMAS_VENTAS)) if sale_line.MONTOS_EN_OTRAS_FORMAS_VENTAS else ""
			lines.append(ln)

		for line in lines:
			sale_file.write(line + "\n")

		sale_file.close()
		sale_file = open(sale_path, 'rb')
		sale_binary = base64.b64encode(sale_file.read())
		report_name = 'DGII_F_607_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
		self.write({'sale_binary': sale_binary, 'sale_filename': report_name})

		''' ************************ 606 TXT REPORT ******************************** '''
		pruchase_path = '/tmp/606{}.txt'.format(company_fiscal_identificacion)
		purchase_file = open(pruchase_path, 'w')
		lines = []

		CANTIDAD_REGISTRO = len(self.purchase_report)

		periodo = str(year) + str(month).zfill(2)
		data = ('606', str(company_fiscal_identificacion), periodo, str(CANTIDAD_REGISTRO))
		header = pipe.join(data)
		lines.append(header)

		for line in self.purchase_report:
			data = []
			data.append(line.RNC_CEDULA)
			data.append(line.TIPO_IDENTIFICACION)
			data.append(line.TIPO_BIENES_SERVICIOS_COMPRADOS)
			data.append(line.NUMERO_COMPROBANTE_FISCAL)
			data.append(line.NUMERO_COMPROBANTE_MODIFICADO or "".rjust(19))
			data.append(line.FECHA_COMPROBANTE.replace("-", "") if line.FECHA_COMPROBANTE else "".rjust(8))
			data.append(line.FECHA_PAGO.replace("-", "") if line.FECHA_PAGO else "".rjust(8))
			data.append(str(abs(line.MONTO_FACTURADO_SERVICIOS)).zfill(12))
			data.append(str(abs(line.MONTO_FACTURADO_BIENES)).zfill(12))
			data.append(str(abs(line.MONTO_FACTURADO)).zfill(12))  # the total
			data.append(str(abs(line.ITBIS_FACTURADO_TOTAL)).zfill(12))
			data.append(str(abs(line.ITBIS_RETENIDO)).zfill(12))
			data.append(str(abs(line.ITBIS_SUJETO_PROPORCIONALIDAD)).zfill(12))
			data.append(str(abs(line.ITBIS_LLEVADO_ALCOSTO)).zfill(12))
			data.append(str(abs(line.ITBIS_POR_ADELANTAR)).zfill(12))
			data.append(str(abs(line.ITBIS_PERCIBIDO_COMPRAS)).zfill(12))
			data.append(line.TIPO_RETENCION_ISR or "00")
			data.append(str(abs(line.RETENCION_RENTA)).zfill(12))
			data.append(str(abs(line.ISR_PERCIBIDO_COMPRAS)).zfill(12))
			data.append(str(abs(line.IMPUESTO_ISC)).zfill(12))
			data.append(str(abs(line.IMPUESTOS_OTROS)).zfill(12))
			data.append(str(abs(line.MONTO_PROPINA_LEGAL)).zfill(12))
			data.append(line.FORMA_PAGO)
			lines.append(pipe.join(data))

		for line in lines:
			purchase_file.write(line + "\n")

		purchase_file.close()
		purchase_file = open(pruchase_path, 'rb')
		purchase_binary = base64.b64encode(purchase_file.read())
		purchase_filename = 'DGII_F_606_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year),
															str(month).zfill(2))
		self.write({'purchase_binary': purchase_binary, 'purchase_filename': purchase_filename})

		''' ************************ 608 TXT REPORT ******************************** '''
		path = '/tmp/608{}.txt'.format(company_fiscal_identificacion)
		file = open(path, 'w')
		lines = []

		header = "608"
		header += company_fiscal_identificacion.zfill(11)
		header += str(year)
		header += str(month).zfill(2)
		lines.append(header)

		for line in self.cancel_report:
			ln = ""
			ln += line.NUMERO_COMPROBANTE_FISCAL
			ln += line.FECHA_COMPROBANTE and line.FECHA_COMPROBANTE.replace("-", "") or ""
			ln += "{}".format(line.TIPO_ANULACION).zfill(2)
			lines.append(ln)

		for line in lines:
			file.write(line + "\n")

		file.close()
		file = open(path, 'rb')
		report = base64.b64encode(file.read())
		report_name = 'DGII_608_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
		self.write({'cancel_binary': report, 'cancel_filename': report_name})

	'''
		DEPRECATED! OLD DGII WAY
	'''

	def generate_txt_files(self):
		company_fiscal_identificacion = re.sub("[^0-9]", "", self.company_id.vat)
		period = self.name.split("/")
		month = period[0]
		year = period[1]

		''' ************************ 607 TXT REPORT ******************************** '''
		sale_path = '/tmp/607{}.txt'.format(company_fiscal_identificacion)
		sale_file = open(sale_path, 'w')

		lines = []

		CANTIDAD_REGISTRO = str(len(self.sale_report)).zfill(12)
		TOTAL_MONTO_FACTURADO_FACTURAS = sum(
			[rec.MONTO_FACTURADO for rec in self.sale_report if rec.NUMERO_COMPROBANTE_MODIFICADO == False])
		TOTAL_MONTO_FACTURADO_NC = sum(
			[rec.MONTO_FACTURADO for rec in self.sale_report if rec.NUMERO_COMPROBANTE_MODIFICADO != False])
		TOTAL_MONTO_FACTURADO = "{:.2f}".format(TOTAL_MONTO_FACTURADO_FACTURAS - TOTAL_MONTO_FACTURADO_NC).zfill(16)

		header = "607"
		header += company_fiscal_identificacion.rjust(11)
		header += str(year)
		header += str(month).zfill(2)
		header += CANTIDAD_REGISTRO
		header += TOTAL_MONTO_FACTURADO
		lines.append(header)

		for sale_line in self.sale_report:
			ln = ""
			ln += sale_line.RNC_CEDULA and sale_line.RNC_CEDULA.rjust(11) or "".rjust(11)
			ln += sale_line.TIPO_IDENTIFICACION
			ln += sale_line.NUMERO_COMPROBANTE_FISCAL.rjust(19)
			ln += sale_line.NUMERO_COMPROBANTE_MODIFICADO or "".rjust(19)
			ln += sale_line.FECHA_COMPROBANTE.replace("-", "")
			ln += "{:.2f}".format(sale_line.ITBIS_FACTURADO).zfill(12)
			ln += "{:.2f}".format(sale_line.MONTO_FACTURADO).zfill(12)
			lines.append(ln)

		for line in lines:
			sale_file.write(line + "\n")

		sale_file.close()
		sale_file = open(sale_path, 'rb')
		sale_binary = base64.b64encode(sale_file.read())
		report_name = 'DGII_607_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
		self.write({'sale_binary': sale_binary, 'sale_filename': report_name})

		''' ************************ 606 TXT REPORT ******************************** '''
		pruchase_path = '/tmp/606{}.txt'.format(company_fiscal_identificacion)
		purchase_file = open(pruchase_path, 'w')
		lines = []

		CANTIDAD_REGISTRO = "{:.2f}".format(len(self.purchase_report)).zfill(12)
		TOTAL_MONTO_FACTURADO_FACTURAS = sum(
			[rec.MONTO_FACTURADO for rec in self.purchase_report if rec.NUMERO_COMPROBANTE_MODIFICADO == False])
		TOTAL_MONTO_FACTURADO_NC = sum(
			[rec.MONTO_FACTURADO for rec in self.purchase_report if rec.NUMERO_COMPROBANTE_MODIFICADO != False])
		TOTAL_MONTO_FACTURADO = "{:.2f}".format(TOTAL_MONTO_FACTURADO_FACTURAS - TOTAL_MONTO_FACTURADO_NC).zfill(16)

		RETENCION_RENTA = "{:.2f}".format(sum([rec.RETENCION_RENTA for rec in self.purchase_report])).zfill(12)

		header = "606"
		header += company_fiscal_identificacion.rjust(11)
		header += str(year)
		header += str(month).zfill(2)
		header += CANTIDAD_REGISTRO
		header += TOTAL_MONTO_FACTURADO
		header += RETENCION_RENTA
		lines.append(header)

		for line in self.purchase_report:
			ln = ""
			ln += line.RNC_CEDULA.rjust(11)
			ln += line.TIPO_IDENTIFICACION
			ln += line.TIPO_BIENES_SERVICIOS_COMPRADOS
			ln += line.NUMERO_COMPROBANTE_FISCAL and line.NUMERO_COMPROBANTE_FISCAL.rjust(19) or "".rjust(19)
			ln += line.NUMERO_COMPROBANTE_MODIFICADO or "".rjust(19)
			ln += line.FECHA_COMPROBANTE.replace("-", "")
			ln += line.FECHA_PAGO.replace("-", "") if line.FECHA_PAGO else "".rjust(8)
			ln += "{:.2f}".format(line.ITBIS_FACTURADO_TOTAL).zfill(12)
			ln += "{:.2f}".format(abs(line.ITBIS_RETENIDO)).zfill(12)
			ln += "{:.2f}".format(line.MONTO_FACTURADO).zfill(12)
			ln += "{:.2f}".format(line.RETENCION_RENTA).zfill(12)
			lines.append(ln)

		for line in lines:
			purchase_file.write(line + "\n")

		purchase_file.close()
		purchase_file = open(pruchase_path, 'rb')
		purchase_binary = base64.b64encode(purchase_file.read())
		purchase_filename = 'DGII_606_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
		self.write({'purchase_binary': purchase_binary, 'purchase_filename': purchase_filename})

		''' ************************ 608 TXT REPORT ******************************** '''
		path = '/tmp/608{}.txt'.format(company_fiscal_identificacion)
		file = open(path, 'w')
		lines = []

		header = "608"
		header += company_fiscal_identificacion.zfill(11)
		header += str(year)
		header += str(month).zfill(2)
		lines.append(header)

		for line in self.cancel_report:
			ln = ""
			ln += line.NUMERO_COMPROBANTE_FISCAL
			ln += line.FECHA_COMPROBANTE and line.FECHA_COMPROBANTE.replace("-", "") or ""
			ln += "{}".format(line.TIPO_ANULACION).zfill(2)
			lines.append(ln)

		for line in lines:
			file.write(line + "\n")

		file.close()
		file = open(path, 'rb')
		report = base64.b64encode(file.read())
		report_name = 'DGII_608_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
		self.write({'cancel_binary': report, 'cancel_filename': report_name})

	def get_type_voucher(self, purchase):

		if len(purchase.NUMERO_COMPROBANTE_FISCAL) == 19:
			return purchase.NUMERO_COMPROBANTE_FISCAL[9:-8]
		else:
			return purchase.NUMERO_COMPROBANTE_FISCAL[1:3]

	'''
		************ Model Properties ************
	'''

	company_id = fields.Many2one('res.company', 'EMPRESA', required=False,
								 default=lambda self: self.env.user.company_id)
	name = fields.Char(string=u"PERÍODO MES/AÑO", required=True, unique=True, index=True)
	positive_balance = fields.Float(u"SALDO A FAVOR ANTERIOR", required=True)

	it_filename = fields.Char()
	it_binary = fields.Binary(string=u"Archivo excel IT-1")

	ir17_filename = fields.Char()
	ir17_binary = fields.Binary(string=u"Archivo excel IR-17")

	# 606
	COMPRAS_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)

	TOTAL_MONTO_FACTURADO = fields.Float(u"Monto compra", compute=_purchase_report_totals)
	MONTO_FACTURADO_SERVICIOS = fields.Float(u"Monto Facturado Servicios", compute=_purchase_report_totals)
	MONTO_FACTURADO_BIENES = fields.Float(u"Monto Facturado Bienes", compute=_purchase_report_totals)

	ITBIS_TOTAL = fields.Float(u"ITBIS Compras", compute=_purchase_report_totals)
	ITBIS_FACTURADO_SERVICIOS = fields.Float(u"ITBIS Facturado Servicios", compute=_purchase_report_totals)
	ITBIS_FACTURADO_BIENES = fields.Float(u"ITBIS Facturado Bienes", compute=_purchase_report_totals)

	TOTAL_MONTO_NC = fields.Float(u"Notas de crédito", compute=_purchase_report_totals)
	ITBIS_TOTAL_NC = fields.Float(u"ITBIS Notas de crédito", compute=_purchase_report_totals)

	TOTAL_MONTO_PAYMENT = fields.Float(u"Total monto facturado", compute=_purchase_report_totals)
	ITBIS_TOTAL_PAYMENT = fields.Float(u"ITBIS Pagado", compute=_purchase_report_totals)

	ITBIS_RETENIDO = fields.Float(u"ITBIS Retenido", compute=_purchase_report_totals)
	RETENCION_RENTA = fields.Float(u"Retención Renta", compute=_purchase_report_totals)

	purchase_report = fields.One2many(u"dgii.report.purchase.line", "dgii_report_id")
	purchase_filename = fields.Char()
	purchase_binary = fields.Binary(string=u"Archivo TXT del Reporte 606")

	# 606 type summary
	currency_id = fields.Many2one(related="company_id.currency_id")

	pcount_01 = fields.Integer(compute=_purchase_report_totals)
	pcount_02 = fields.Integer(compute=_purchase_report_totals)
	pcount_03 = fields.Integer(compute=_purchase_report_totals)
	pcount_04 = fields.Integer(compute=_purchase_report_totals)
	pcount_05 = fields.Integer(compute=_purchase_report_totals)
	pcount_06 = fields.Integer(compute=_purchase_report_totals)
	pcount_07 = fields.Integer(compute=_purchase_report_totals)
	pcount_08 = fields.Integer(compute=_purchase_report_totals)
	pcount_09 = fields.Integer(compute=_purchase_report_totals)
	pcount_10 = fields.Integer(compute=_purchase_report_totals)
	pcount_11 = fields.Integer(compute=_purchase_report_totals)

	pamount_01 = fields.Monetary(compute=_purchase_report_totals)
	pamount_02 = fields.Monetary(compute=_purchase_report_totals)
	pamount_03 = fields.Monetary(compute=_purchase_report_totals)
	pamount_04 = fields.Monetary(compute=_purchase_report_totals)
	pamount_05 = fields.Monetary(compute=_purchase_report_totals)
	pamount_06 = fields.Monetary(compute=_purchase_report_totals)
	pamount_07 = fields.Monetary(compute=_purchase_report_totals)
	pamount_08 = fields.Monetary(compute=_purchase_report_totals)
	pamount_09 = fields.Monetary(compute=_purchase_report_totals)
	pamount_10 = fields.Monetary(compute=_purchase_report_totals)
	pamount_11 = fields.Monetary(compute=_purchase_report_totals)

	# 607
	VENTAS_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)

	SALE_TOTAL_MONTO_FACTURADO = fields.Float(u"Total Facturado", compute=_sale_report_totals)
	SALE_ITBIS_TOTAL = fields.Float(u"ITBIS ventas", compute=_sale_report_totals)

	SALE_TOTAL_MONTO_NC = fields.Float(u"Total Notas de crédito", compute=_sale_report_totals)
	SALE_ITBIS_NC = fields.Float(u"ITBIS Notas de crédito", compute=_sale_report_totals)

	SALE_TOTAL_MONTO_CHARGED = fields.Float(u"Facturado", compute=_sale_report_totals)
	SALE_ITBIS_CHARGED = fields.Float(u"ITBIS Cobrado", compute=_sale_report_totals)
	MONTO_FACTURADO_EXCENTO = fields.Float(u"ITBIS Cobrado", compute=_sale_report_totals)

	sale_filename = fields.Char()
	sale_binary = fields.Binary(string=u"Archivo 607 TXT")

	sale_report = fields.One2many("dgii.report.sale.line", "dgii_report_id")

	# 607 type summary
	count_final = fields.Integer(compute=_sale_report_totals)
	count_fiscal = fields.Integer(compute=_sale_report_totals)
	count_gov = fields.Integer(compute=_sale_report_totals)
	count_special = fields.Integer(compute=_sale_report_totals)
	count_unico = fields.Integer(compute=_sale_report_totals)
	amount_final = fields.Integer(compute=_sale_report_totals)
	amount_fiscal = fields.Integer(compute=_sale_report_totals)
	amount_gov = fields.Integer(compute=_sale_report_totals)
	amount_special = fields.Integer(compute=_sale_report_totals)
	amount_unico = fields.Integer(compute=_sale_report_totals)

	# 608
	CANCEL_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)
	cancel_report = fields.One2many("dgii.cancel.report.line", "dgii_report_id")
	cancel_filename = fields.Char()
	cancel_binary = fields.Binary(string=u"Archivo 608 TXT")

	# 609
	EXTERIOR_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)
	EXTERIOR_TOTAL_MONTO_FACTURADO = fields.Float()
	exterior_report = fields.One2many("dgii.exterior.report.line", "dgii_report_id")
	exterior_filename = fields.Char(u"Total Monto Facturado")
	exterior_binary = fields.Binary(string=u"Archivo 607 TXT")

	state = fields.Selection([('draft', 'Nuevo'), ('error', 'Con errores'), ('done', 'Validado')], default="draft")


class DgiiReportPurchaseLine(models.Model):
	_name = "dgii.report.purchase.line"
	_order = "FECHA_COMPROBANTE, NUMERO_COMPROBANTE_FISCAL ASC"

	def get_str_forma_pago(self, FORMA_PAGO):

		FORMA_PAGO_STR = FORMA_PAGO

		if FORMA_PAGO == '01':
			FORMA_PAGO_STR = 'Efectivo (01)'
		elif FORMA_PAGO == '02':
			FORMA_PAGO_STR = 'Ch/Trans/Dep. (02)'
		elif FORMA_PAGO == '03':
			FORMA_PAGO_STR = 'TC/TD (03)'
		elif FORMA_PAGO == '04':
			FORMA_PAGO_STR = 'Compra cred. (04)'
		elif FORMA_PAGO == '05':
			FORMA_PAGO_STR = 'Permuta (05)'
		elif FORMA_PAGO == '06':
			FORMA_PAGO_STR = 'Nota cred. (06)'
		elif FORMA_PAGO == '07':
			FORMA_PAGO_STR = 'Mixto (07)'

		return FORMA_PAGO_STR

	def _get_str(self):
		ISR_RETENTION_TYPE = {
			False: '',
			'01': 'Alquileres (01)',
			'02': 'Honorarios por servicios (02)',
			'03': 'Otras rentas (03)',
			'04': 'Otras rentas (rentas presuntas) (04)',
			'05': 'Intereses pagados a personas jurídicas residentes (05)',
			'06': 'Intereses pagados a personas físicas residentes (06)',
			'07': 'Retención por proveedores del Estado (07)',
			'08': 'Juegos telefónicos (08)'
		}

		for rec in self:
			rec.TIPO_IDENTIFICACION_STR = "RNC (1)" if rec.TIPO_IDENTIFICACION == '1' else "C.I. (2)"
			rec.FORMA_PAGO_STR = self.get_str_forma_pago(rec.FORMA_PAGO)
			rec.TIPO_RETENCION_ISR_STR = ISR_RETENTION_TYPE[rec.TIPO_RETENCION_ISR]

	dgii_report_id = fields.Many2one("dgii.report")
	LINE = fields.Integer("Line")
	TIPO_BIENES_SERVICIOS_COMPRADOS = fields.Char(u"3 - Tipo Bienes/Servicios", size=2)
	RNC_CEDULA = fields.Char(u"1 - RNC", size=11)
	TIPO_IDENTIFICACION = fields.Char(u"2 - Tipo Identificación", size=1)
	NUMERO_COMPROBANTE_FISCAL = fields.Char(u"4 - NCF", size=19)
	NUMERO_COMPROBANTE_MODIFICADO = fields.Char(u"5 - NCF Modificado", size=19)
	FECHA_COMPROBANTE = fields.Date(u"6 - Fecha NCF")
	FECHA_PAGO = fields.Date(u"7 - Fecha Pago")
	MONTO_FACTURADO_SERVICIOS = fields.Float(u"8 - Monto Facturado (Servicios)")
	MONTO_FACTURADO_BIENES = fields.Float(u"9 - Monto Facturado (Bienes)")
	MONTO_FACTURADO = fields.Float(u"10 - Monto Facturado (Total)")
	ITBIS_FACTURADO_TOTAL = fields.Float(u"11 - ITBIS Facturado (Total)")
	ITBIS_FACTURADO_BIENES = fields.Float(u"ITBIS Facturado (Bienes)")
	ITBIS_FACTURADO_SERVICIOS = fields.Float(u"ITBIS Facturado (Servicios)")
	ITBIS_RETENIDO = fields.Float(u"12 - ITBIS Retenido")
	ITBIS_SUJETO_PROPORCIONALIDAD = fields.Float(u"13 - ITBIS sujeto a Proporcionalidad (Art. 349)")
	ITBIS_LLEVADO_ALCOSTO = fields.Float(u"14 - ITBIS llevado al Costo")
	ITBIS_POR_ADELANTAR = fields.Float(u"15 - ITBIS por Adelantar")
	ITBIS_PERCIBIDO_COMPRAS = fields.Float(u"16 - ITBIS percibido en compras")
	TIPO_RETENCION_ISR = fields.Char(u"17 - Tipo de Retención en ISR", size=2)
	RETENCION_RENTA = fields.Float(u"18 - Monto Retención Renta")
	ISR_PERCIBIDO_COMPRAS = fields.Float(u"19 - ISR Percibido en compras")
	IMPUESTO_ISC = fields.Float(u"20 - Impuesto Selectivo al Consumo")
	IMPUESTOS_OTROS = fields.Float(u"21 - Otros Impuesto/Tasas")
	MONTO_PROPINA_LEGAL = fields.Float(u"22 - Monto Propina Legal")
	FORMA_PAGO = fields.Char(u"23 - Forma de Pago", size=2)

	invoice_id = fields.Many2one("account.invoice", "NCF")
	number = fields.Char(related="invoice_id.number", string=" NCF")
	inv_partner = fields.Many2one("res.partner", related="invoice_id.partner_id", string="1 - Proveedor")
	affected_nvoice_id = fields.Many2one("account.invoice", "Relacionado NCF Modificado")
	nc = fields.Boolean()

	TIPO_IDENTIFICACION_STR = fields.Char(u"2 - Tipo Identificación", compute=_get_str)
	FORMA_PAGO_STR = fields.Char(u"23 - Forma de Pago", compute=_get_str, size=20)
	TIPO_RETENCION_ISR_STR = fields.Char(u"17 - Tipo de Retención en ISR", compute=_get_str, size=30)


class DgiiReportSaleLine(models.Model):
	_name = "dgii.report.sale.line"
	_order = "FECHA_COMPROBANTE, NUMERO_COMPROBANTE_FISCAL ASC"

	def _get_str(self):
		INCOME_TYPE = {
			1: 'Ingresos x operaciones (1)',
			2: 'Ingresos Financieros (2)',
			3: 'Ingresos Extraordinarios (3)',
			4: 'Ingresos por Arrendamientos (4)',
			5: 'Ingresos por Venta de Activo Depreciable (5)',
			6: 'Otros Ingresos (5)'
		}

		for rec in self:
			rec.TIPO_DE_INGRESO_STR = INCOME_TYPE[rec.TIPO_DE_INGRESO]

	dgii_report_id = fields.Many2one("dgii.report")
	LINE = fields.Integer("Line")
	RNC_CEDULA = fields.Char(u"1 - RNC", size=11)
	TIPO_IDENTIFICACION = fields.Char(u"2 - Tipo Identificación", size=1)
	NUMERO_COMPROBANTE_FISCAL = fields.Char("3 - NCF", size=19)
	NUMERO_COMPROBANTE_MODIFICADO = fields.Char(u"4 - NCF Modificado", size=19)
	TIPO_DE_INGRESO = fields.Integer("5 - Tipo de Ingreso")  # new
	FECHA_COMPROBANTE = fields.Date(u"6 - Fecha NCF")
	FECHA_RETENCION = fields.Date(u"7 - Fecha Retención")  # new
	MONTO_FACTURADO = fields.Float(u"8 - Monto Facturado")
	ITBIS_FACTURADO = fields.Float(u"9 - ITBIS Facturado")
	ITBIS_RETENIDO_POR_TERCEROS = fields.Float(u"10 - ITBIS Retenido")  # new
	ITBIS_PERCIBIDO = fields.Float(u"11 - ITBIS Percibido")  # new
	RETENCION_RENTA_POR_TERCEROS = fields.Float(u"12 - Retención Renta")  # new
	ISR_PERCIBIDO = fields.Float(u"13 - ISR Percibido")  # new
	IMPUESTO_ISC = fields.Float(u"14 - ISC")  # new
	IMPUESTOS_OTROS = fields.Float(u"15 - OTROS IMP.")  # new
	MONTO_PROPINA_LEGAL = fields.Float(u"16 - Prop. Legal")  # new
	MONTOS_PAGADOS_EFECTIVO = fields.Float(u"17 - Efectivo")  # new
	MONTOS_PAGADOS_BANCO = fields.Float(u"18 - CH/TRANS/DEPT")  # new
	MONTOS_PAGADOS_TARJETAS = fields.Float(u"19 - Tarjetas")  # new
	MONTOS_A_CREDITO = fields.Float(u"20 - A Crédito")  # new
	MONTOS_EN_BONOS_O_CERTIFICADOS_REGALOS = fields.Float(u"21 - Bonos/Regalos")  # new
	MONTOS_EN_PERMUTA = fields.Float(u"22 - Permuta")  # new
	MONTOS_EN_OTRAS_FORMAS_VENTAS = fields.Float(u"23 - Permuta")  # new

	MONTO_FACTURADO_EXCENTO = fields.Float(u"Monto Facturado Exento")  # NO USADO EN 607

	invoice_id = fields.Many2one("account.invoice", "NCF")
	currency_id = fields.Many2one('res.currency', string='Currency', related="invoice_id.currency_id",
								  required=True, readonly=True, states={'draft': [('readonly', False)]},
								  track_visibility='always')  # todo validate to remove, IT IS NOT IN THE DB.

	number = fields.Char(related="invoice_id.number", string=" NCF")  # todo validate to remove, IT IS NOT IN THE DB.
	inv_partner = fields.Many2one("res.partner", related="invoice_id.partner_id", string="Cliente")
	affected_nvoice_id = fields.Many2one("account.invoice", "NCF Modificado")
	nc = fields.Boolean()

	TIPO_DE_INGRESO_STR = fields.Char(u"5 - Tipo de Ingreso", compute=_get_str, size=50)


class DgiiCancelReportline(models.Model):
	_name = "dgii.cancel.report.line"

	dgii_report_id = fields.Many2one("dgii.report")
	LINE = fields.Integer("Linea")
	NUMERO_COMPROBANTE_FISCAL = fields.Char("NCF", size=19)
	FECHA_COMPROBANTE = fields.Date("Fecha")
	TIPO_ANULACION = fields.Char(u"Tipo de anulación", size=2)
	invoice_id = fields.Many2one("account.invoice", "Factura")


class DgiiExteriorReportline(models.Model):
	_name = "dgii.exterior.report.line"

	dgii_report_id = fields.Many2one("dgii.report")
	LINE = fields.Integer("Linea")
	TIPO_BIENES_SERVICIOS_COMPRADOS = fields.Char("Tipo", size=2)
	FECHA_COMPROBANTE = fields.Date("Fecha")
	FECHA_PAGO = fields.Date("Pagado")
	RETENCION_RENTA = fields.Float(u"Retención Renta")
	MONTO_FACTURADO = fields.Float("Monto Facturado")
	invoice_id = fields.Many2one("account.invoice", "Factura")
